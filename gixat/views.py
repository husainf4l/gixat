from django.shortcuts import render, redirect
from django.contrib.auth import login
from django.contrib.auth.views import LoginView
from django.contrib import messages
from django.urls import reverse_lazy
from django.db import transaction
from django.db.models import Count, Sum, Q, F, Avg
from django.utils import timezone
from datetime import timedelta
from .forms import CustomLoginForm, CustomSignupForm, OrganizationRegistrationForm, AdminUserCreationForm, ClientRequestForm, InspectionForm, VehicleForm, InventoryItemForm, SessionForm, ProfileForm, OrganizationForm, PasswordChangeForm, NotificationPreferencesForm, SystemSettingsForm
from .models import Organization, UserProfile, Client, Car, Session, JobCard, Inventory, InventoryTransaction, Inspection, Notification
import csv
from django.http import HttpResponse
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def home(request):
    return render(request, 'home.html')


def about(request):
    return render(request, 'about.html')


def dashboard(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    # Get current date and time
    now = timezone.now()
    today = now.date()
    this_month = today.replace(day=1)
    this_year = today.replace(month=1, day=1)
    
    # Quick Overview KPIs
    cars_in_garage = Session.objects.filter(
        organization=organization,
        status__in=['scheduled', 'in_progress']
    ).count()
    
    pending_requests = Session.objects.filter(
        organization=organization,
        status='scheduled'
    ).count()
    
    inspections_in_progress = Inspection.objects.filter(
        organization=organization,
        status='in_progress'
    ).count()
    
    jobs_in_progress = JobCard.objects.filter(
        session__organization=organization,
        status='in_progress'
    ).count()
    
    inventory_alerts = Inventory.objects.filter(
        organization=organization,
        is_active=True,
        quantity__lte=F('min_quantity')
    ).count()
    
    # Revenue calculations
    revenue_today = Session.objects.filter(
        organization=organization,
        status='completed',
        actual_end_time__date=today
    ).aggregate(total=Sum('actual_cost'))['total'] or 0
    
    revenue_this_month = Session.objects.filter(
        organization=organization,
        status='completed',
        actual_end_time__date__gte=this_month
    ).aggregate(total=Sum('actual_cost'))['total'] or 0
    
    revenue_this_year = Session.objects.filter(
        organization=organization,
        status='completed',
        actual_end_time__date__gte=this_year
    ).aggregate(total=Sum('actual_cost'))['total'] or 0
    
    # Active Cars List
    active_sessions = Session.objects.filter(
        organization=organization,
        status__in=['scheduled', 'in_progress']
    ).select_related('car', 'car__client', 'technician__user').order_by('-scheduled_date')[:10]
    
    # Work Orders Board
    new_requests = Session.objects.filter(organization=organization, status='scheduled')
    approved_assigned = Session.objects.filter(organization=organization, status='in_progress')
    completed_sessions = Session.objects.filter(organization=organization, status='completed')[:5]
    delivered = Session.objects.filter(
        organization=organization, 
        status='completed',
        actual_end_time__date__gte=today - timedelta(days=7)
    )[:5]
    
    # Recent Inspections
    recent_inspections = Inspection.objects.filter(
        organization=organization
    ).select_related('car', 'car__client', 'inspector__user').order_by('-created_at')[:5]
    
    # Inventory Snapshot
    low_stock_items = Inventory.objects.filter(
        organization=organization,
        is_active=True,
        quantity__lte=F('min_quantity')
    ).order_by('quantity')[:5]
    
    # Notifications/Alerts
    notifications = Notification.objects.filter(
        organization=organization,
        user=request.user
    ).order_by('-created_at')[:10]
    
    # Reports/Analytics
    cars_processed_this_week = Session.objects.filter(
        organization=organization,
        status='completed',
        actual_end_time__date__gte=today - timedelta(days=7)
    ).count()
    
    cars_processed_this_month = Session.objects.filter(
        organization=organization,
        status='completed',
        actual_end_time__date__gte=this_month
    ).count()
    
    # Services breakdown (most common jobs)
    services_breakdown = JobCard.objects.filter(
        session__organization=organization,
        status='completed'
    ).values('title').annotate(count=Count('id')).order_by('-count')[:5]
    
    # Staff productivity
    staff_productivity = UserProfile.objects.filter(
        organization=organization,
        role__in=['technician', 'manager']
    ).annotate(
        completed_jobs=Count('session', filter=Q(session__status='completed'))
    ).order_by('-completed_jobs')[:5]
    
    context = {
        # KPIs
        'cars_in_garage': cars_in_garage,
        'pending_requests': pending_requests,
        'inspections_in_progress': inspections_in_progress,
        'jobs_in_progress': jobs_in_progress,
        'inventory_alerts': inventory_alerts,
        'revenue_today': revenue_today,
        'revenue_this_month': revenue_this_month,
        'revenue_this_year': revenue_this_year,
        
        # Active Cars
        'active_sessions': active_sessions,
        
        # Work Orders
        'new_requests': new_requests,
        'approved_assigned': approved_assigned,
        'completed_sessions': completed_sessions,
        'delivered': delivered,
        
        # Recent Inspections
        'recent_inspections': recent_inspections,
        
        # Inventory
        'low_stock_items': low_stock_items,
        
        # Notifications
        'notifications': notifications,
        
        # Analytics
        'cars_processed_this_week': cars_processed_this_week,
        'cars_processed_this_month': cars_processed_this_month,
        'services_breakdown': services_breakdown,
        'staff_productivity': staff_productivity,
    }
    
    return render(request, 'dashboard.html', context)


def appointments(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    # Get all sessions/appointments
    sessions = Session.objects.filter(
        organization=organization
    ).select_related('car', 'car__client', 'technician__user').order_by('-scheduled_date')
    
    # Filter by status if provided
    status_filter = request.GET.get('status')
    if status_filter:
        sessions = sessions.filter(status=status_filter)
    
    context = {
        'sessions': sessions,
        'status_filter': status_filter,
    }
    
    return render(request, 'appointments.html', context)


def services(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    # Get all job cards/services
    job_cards = JobCard.objects.filter(
        session__organization=organization
    ).select_related('session', 'session__car', 'session__car__client', 'assigned_technician__user').order_by('-created_at')
    
    # Filter by status if provided
    status_filter = request.GET.get('status')
    if status_filter:
        job_cards = job_cards.filter(status=status_filter)
    
    # Calculate statistics
    total_services = job_cards.count()
    pending_services = job_cards.filter(status='pending').count()
    in_progress_services = job_cards.filter(status='in_progress').count()
    completed_services = job_cards.filter(status='completed').count()
    
    # Calculate total revenue from completed services
    total_revenue = job_cards.filter(status='completed').aggregate(
        total=Sum(F('labor_cost') + F('parts_cost'))
    )['total'] or 0
    
    # Get recent activity (last 7 days)
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    recent_date = timezone.now() - timedelta(days=7)
    recent_services = job_cards.filter(created_at__gte=recent_date).count()
    
    context = {
        'job_cards': job_cards,
        'status_filter': status_filter,
        'total_services': total_services,
        'pending_services': pending_services,
        'in_progress_services': in_progress_services,
        'completed_services': completed_services,
        'total_revenue': total_revenue,
        'recent_services': recent_services,
    }
    
    return render(request, 'services.html', context)


def services_in_progress(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    # Get all in-progress job cards/services
    job_cards = JobCard.objects.filter(
        session__organization=organization,
        status='in_progress'
    ).select_related('session', 'session__car', 'session__car__client', 'assigned_technician__user').order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        job_cards = job_cards.filter(
            Q(job_number__icontains=search_query) |
            Q(title__icontains=search_query) |
            Q(session__car__make__icontains=search_query) |
            Q(session__car__model__icontains=search_query) |
            Q(session__car__license_plate__icontains=search_query) |
            Q(session__car__client__full_name__icontains=search_query) |
            Q(assigned_technician__user__first_name__icontains=search_query) |
            Q(assigned_technician__user__last_name__icontains=search_query)
        )
    
    # Calculate statistics
    total_services = job_cards.count()
    
    # Calculate total estimated value of in-progress work
    total_estimated_value = job_cards.aggregate(
        total=Sum(F('labor_cost') + F('parts_cost'))
    )['total'] or 0
    
    # Get technician workload
    technician_workload = job_cards.values(
        'assigned_technician__user__first_name',
        'assigned_technician__user__last_name'
    ).annotate(
        job_count=Count('id'),
        total_hours=Sum('estimated_hours'),
        total_value=Sum(F('labor_cost') + F('parts_cost'))
    ).order_by('-job_count')
    
    # Get priority breakdown
    priority_breakdown = job_cards.values('priority').annotate(
        count=Count('id')
    ).order_by('-count')
    
    # Get recent activity (jobs started in last 7 days)
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    recent_date = timezone.now() - timedelta(days=7)
    recent_started = job_cards.filter(started_at__gte=recent_date).count()
    
    context = {
        'job_cards': job_cards,
        'search_query': search_query,
        'total_services': total_services,
        'total_estimated_value': total_estimated_value,
        'technician_workload': technician_workload,
        'priority_breakdown': priority_breakdown,
        'recent_started': recent_started,
        'page_title': 'Services In Progress',
        'status_filter': 'in_progress',
    }
    
    return render(request, 'services_in_progress.html', context)


def services_completed(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    # Get all completed job cards/services
    job_cards = JobCard.objects.filter(
        session__organization=organization,
        status='completed'
    ).select_related('session', 'session__car', 'session__car__client', 'assigned_technician__user').order_by('-completed_at')
    
    # Date filtering
    date_filter = request.GET.get('date_filter')
    if date_filter:
        from datetime import datetime, timedelta
        from django.utils import timezone
        
        today = timezone.now().date()
        if date_filter == 'today':
            job_cards = job_cards.filter(completed_at__date=today)
        elif date_filter == 'week':
            week_start = today - timedelta(days=today.weekday())
            job_cards = job_cards.filter(completed_at__date__gte=week_start)
        elif date_filter == 'month':
            month_start = today.replace(day=1)
            job_cards = job_cards.filter(completed_at__date__gte=month_start)
        elif date_filter == 'year':
            year_start = today.replace(month=1, day=1)
            job_cards = job_cards.filter(completed_at__date__gte=year_start)
    
    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        job_cards = job_cards.filter(
            Q(job_number__icontains=search_query) |
            Q(title__icontains=search_query) |
            Q(session__car__make__icontains=search_query) |
            Q(session__car__model__icontains=search_query) |
            Q(session__car__license_plate__icontains=search_query) |
            Q(session__car__client__full_name__icontains=search_query) |
            Q(assigned_technician__user__first_name__icontains=search_query) |
            Q(assigned_technician__user__last_name__icontains=search_query)
        )
    
    # Calculate statistics
    total_services = job_cards.count()
    
    # Calculate total revenue from completed services
    total_revenue = job_cards.aggregate(
        total=Sum(F('labor_cost') + F('parts_cost'))
    )['total'] or 0
    
    # Calculate average job value
    avg_job_value = total_revenue / total_services if total_services > 0 else 0
    
    # Calculate average completion time
    avg_hours = job_cards.filter(actual_hours__isnull=False).aggregate(
        avg=Avg('actual_hours')
    )['avg'] or 0
    
    # Get top performing technicians
    top_technicians = job_cards.values(
        'assigned_technician__user__first_name',
        'assigned_technician__user__last_name'
    ).annotate(
        job_count=Count('id'),
        total_revenue=Sum(F('labor_cost') + F('parts_cost')),
        avg_hours=Avg('actual_hours')
    ).order_by('-total_revenue')[:5]
    
    # Get service type breakdown
    service_breakdown = job_cards.values('title').annotate(
        count=Count('id'),
        total_revenue=Sum(F('labor_cost') + F('parts_cost'))
    ).order_by('-count')[:10]
    
    # Add average value to each service breakdown item
    for service in service_breakdown:
        service['avg_value'] = service['total_revenue'] / service['count'] if service['count'] > 0 else 0
    
    # Get completion trends (last 30 days)
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    completion_trends = []
    for i in range(29, -1, -1):
        date = timezone.now().date() - timedelta(days=i)
        count = job_cards.filter(completed_at__date=date).count()
        revenue = job_cards.filter(completed_at__date=date).aggregate(
            total=Sum(F('labor_cost') + F('parts_cost'))
        )['total'] or 0
        completion_trends.append({
            'date': date,
            'count': count,
            'revenue': revenue
        })
    
    context = {
        'job_cards': job_cards,
        'search_query': search_query,
        'date_filter': date_filter,
        'total_services': total_services,
        'total_revenue': total_revenue,
        'avg_job_value': avg_job_value,
        'avg_hours': avg_hours,
        'top_technicians': top_technicians,
        'service_breakdown': service_breakdown,
        'completion_trends': completion_trends,
        'page_title': 'Completed Services',
        'status_filter': 'completed',
    }
    
    return render(request, 'services_completed.html', context)


def inventory(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    # Get all inventory items
    inventory_items = Inventory.objects.filter(
        organization=organization
    ).order_by('name')
    
    # Filter by category if provided
    category_filter = request.GET.get('category')
    if category_filter:
        inventory_items = inventory_items.filter(category=category_filter)
    
    # Get low stock items
    low_stock_items = Inventory.objects.filter(
        organization=organization,
        is_active=True,
        quantity__lte=F('min_quantity')
    ).order_by('quantity')
    
    context = {
        'inventory_items': inventory_items,
        'low_stock_items': low_stock_items,
        'category_filter': category_filter,
        'categories': Inventory.CATEGORY_CHOICES,
    }
    
    return render(request, 'inventory.html', context)


def reports(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    # Get current date and time
    now = timezone.now()
    today = now.date()
    this_month = today.replace(day=1)
    this_year = today.replace(month=1, day=1)
    
    # Revenue reports
    revenue_today = Session.objects.filter(
        organization=organization,
        status='completed',
        actual_end_time__date=today
    ).aggregate(total=Sum('actual_cost'))['total'] or 0
    
    revenue_this_month = Session.objects.filter(
        organization=organization,
        status='completed',
        actual_end_time__date__gte=this_month
    ).aggregate(total=Sum('actual_cost'))['total'] or 0
    
    revenue_this_year = Session.objects.filter(
        organization=organization,
        status='completed',
        actual_end_time__date__gte=this_year
    ).aggregate(total=Sum('actual_cost'))['total'] or 0
    
    # Service statistics
    total_sessions = Session.objects.filter(organization=organization).count()
    completed_sessions = Session.objects.filter(organization=organization, status='completed').count()
    completion_rate = (completed_sessions / total_sessions * 100) if total_sessions > 0 else 0
    
    # Most popular services
    popular_services = JobCard.objects.filter(
        session__organization=organization,
        status='completed'
    ).values('title').annotate(
        count=Count('id'),
        revenue=Sum(F('parts_cost') + F('labor_cost'))
    ).order_by('-count')[:10]
    
    # Technician performance
    technician_performance = UserProfile.objects.filter(
        organization=organization,
        role__in=['technician', 'manager']
    ).annotate(
        sessions_completed=Count('session', filter=Q(session__status='completed')),
        total_revenue=Sum('session__actual_cost', filter=Q(session__status='completed'))
    ).order_by('-sessions_completed')
    
    # Monthly revenue trend (last 12 months)
    monthly_revenue = []
    for i in range(11, -1, -1):
        month_start = (now - timedelta(days=30*i)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        revenue = Session.objects.filter(
            organization=organization,
            status='completed',
            actual_end_time__date__gte=month_start,
            actual_end_time__date__lte=month_end
        ).aggregate(total=Sum('actual_cost'))['total'] or 0
        monthly_revenue.append({
            'month': month_start.strftime('%B %Y'),
            'revenue': revenue
        })
    
    context = {
        'revenue_today': revenue_today,
        'revenue_this_month': revenue_this_month,
        'revenue_this_year': revenue_this_year,
        'total_sessions': total_sessions,
        'completed_sessions': completed_sessions,
        'completion_rate': completion_rate,
        'popular_services': popular_services,
        'technician_performance': technician_performance,
        'monthly_revenue': monthly_revenue,
    }
    
    return render(request, 'reports.html', context)


def repair_sessions(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    # Get all sessions for the organization with related data (newest first)
    sessions = Session.objects.filter(
        organization=organization
    ).select_related('car', 'car__client', 'technician__user').order_by('-created_at', '-id')
    
    # Filter by status if provided
    status_filter = request.GET.get('status')
    if status_filter:
        sessions = sessions.filter(status=status_filter)
    
    # Search functionality
    search_query = request.GET.get('search')
    if search_query:
        sessions = sessions.filter(
            Q(session_number__icontains=search_query) |
            Q(car__make__icontains=search_query) |
            Q(car__model__icontains=search_query) |
            Q(car__license_plate__icontains=search_query) |
            Q(car__client__full_name__icontains=search_query)
        )
    
    context = {
        'sessions': sessions,
        'status_filter': status_filter,
        'search_query': search_query,
        'total_sessions': sessions.count(),
        'active_sessions': sessions.filter(status__in=['scheduled', 'in_progress']).count(),
        'completed_sessions': sessions.filter(status='completed').count(),
    }
    
    return render(request, 'repair_sessions.html', context)


def signup_choice_view(request):
    return render(request, 'signup_choice.html')


def register_organization_view(request):
    if request.method == 'POST':
        org_form = OrganizationRegistrationForm(request.POST)
        user_form = AdminUserCreationForm(request.POST)
        
        if org_form.is_valid() and user_form.is_valid():
            try:
                with transaction.atomic():
                    # Create organization
                    organization = org_form.save()
                    
                    # Create admin user with profile
                    user = user_form.save(commit=True, organization=organization)
                    
                    # Log the user in
                    login(request, user)
                    
                    messages.success(request, f'Organization "{organization.name}" created successfully! Welcome, {user.get_full_name()}!')
                    return redirect('home')
                    
            except Exception as e:
                messages.error(request, 'An error occurred while creating your organization. Please try again.')
    else:
        org_form = OrganizationRegistrationForm()
        user_form = AdminUserCreationForm()
    
    return render(request, 'register_organization.html', {
        'form': org_form,
        'user_form': user_form
    })


def signup_view(request):
    if request.method == 'POST':
        form = CustomSignupForm(request.POST)
        if form.is_valid():
            user = form.save()
            username = form.cleaned_data.get('username')
            messages.success(request, f'Account created for {username}! You can now login.')
            return redirect('login')
    else:
        form = CustomSignupForm()
    
    return render(request, 'signup.html', {'form': form})


def settings(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    # Initialize forms
    profile_form = ProfileForm(instance=request.user, user_profile=user_profile)
    organization_form = OrganizationForm(instance=organization)
    password_form = PasswordChangeForm(user=request.user)
    preferences_form = NotificationPreferencesForm(initial={
        'email_new_appointments': getattr(user_profile, 'email_notifications', True),
        'email_completed_services': getattr(user_profile, 'email_notifications', True),
        'email_inventory_alerts': getattr(user_profile, 'email_notifications', True),
        'sms_urgent_matters': getattr(user_profile, 'sms_notifications', False),
    })
    system_form = SystemSettingsForm(initial={
        'timezone': getattr(organization, 'timezone', 'UTC'),
        'currency': getattr(organization, 'currency', 'USD'),
        'auto_backup': getattr(organization, 'auto_backup', False),
    })
    
    if request.method == 'POST':
        # Handle profile update
        if 'save_profile' in request.POST:
            profile_form = ProfileForm(request.POST, instance=request.user, user_profile=user_profile)
            if profile_form.is_valid():
                profile_form.save()
                messages.success(request, 'Profile updated successfully!')
                return redirect('settings')
        
        # Handle organization update
        elif 'save_organization' in request.POST:
            organization_form = OrganizationForm(request.POST, instance=organization)
            if organization_form.is_valid():
                organization_form.save()
                messages.success(request, 'Organization settings updated successfully!')
                return redirect('settings')
        
        # Handle password change
        elif 'change_password' in request.POST:
            password_form = PasswordChangeForm(request.POST, user=request.user)
            if password_form.is_valid():
                request.user.set_password(password_form.cleaned_data['new_password'])
                request.user.save()
                messages.success(request, 'Password changed successfully! Please log in again.')
                return redirect('login')
        
        # Handle preferences update
        elif 'save_preferences' in request.POST:
            preferences_form = NotificationPreferencesForm(request.POST)
            if preferences_form.is_valid():
                # Save preferences to user profile or settings
                user_profile.email_notifications = preferences_form.cleaned_data.get('email_new_appointments', False) or preferences_form.cleaned_data.get('email_completed_services', False) or preferences_form.cleaned_data.get('email_inventory_alerts', False)
                user_profile.sms_notifications = preferences_form.cleaned_data.get('sms_urgent_matters', False)
                user_profile.save()
                messages.success(request, 'Notification preferences updated successfully!')
                return redirect('settings')
        
        # Handle system settings update
        elif 'save_settings' in request.POST:
            system_form = SystemSettingsForm(request.POST)
            if system_form.is_valid():
                # Save system settings to organization or user profile
                organization.timezone = system_form.cleaned_data.get('timezone', 'UTC')
                organization.currency = system_form.cleaned_data.get('currency', 'USD')
                organization.auto_backup = system_form.cleaned_data.get('auto_backup', False)
                organization.save()
                messages.success(request, 'System settings updated successfully!')
                return redirect('settings')
        
        # Handle account deletion
        elif 'delete_account' in request.POST:
            # Mark user as inactive instead of deleting
            request.user.is_active = False
            request.user.save()
            user_profile.is_active = False
            user_profile.save()
            messages.success(request, 'Account deactivated successfully.')
            return redirect('home')
    
    context = {
        'profile_form': profile_form,
        'organization_form': organization_form,
        'password_form': password_form,
        'preferences_form': preferences_form,
        'system_form': system_form,
        'user_profile': user_profile,
        'organization': organization,
    }
    
    return render(request, 'settings.html', context)


def new_client_request(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    if request.method == 'POST':
        form = ClientRequestForm(request.POST, organization=organization)
        if form.is_valid():
            try:
                with transaction.atomic():
                    # Handle client creation/selection
                    if form.cleaned_data.get('existing_client'):
                        client = form.cleaned_data['existing_client']
                    else:
                        # Create new client
                        client = Client.objects.create(
                            organization=organization,
                            first_name=form.cleaned_data['first_name'],
                            last_name=form.cleaned_data['last_name'],
                            email=form.cleaned_data.get('email'),
                            phone=form.cleaned_data.get('phone')
                        )
                    
                    # Handle car creation/selection
                    if form.cleaned_data.get('existing_car'):
                        car = form.cleaned_data['existing_car']
                    else:
                        # Create new car
                        car = Car.objects.create(
                            organization=organization,
                            client=client,
                            make=form.cleaned_data['make'],
                            model=form.cleaned_data['model'],
                            year=form.cleaned_data['year'],
                            license_plate=form.cleaned_data['license_plate']
                        )
                    
                    # Create session
                    session = form.save(commit=False)
                    session.organization = organization
                    session.car = car
                    session.technician = user_profile  # Assign to current user or could be selected
                    session.save()
                    
                    messages.success(request, f'Client request created successfully! Session #{session.session_number} has been scheduled.')
                    return redirect('repair_sessions')
                    
            except Exception as e:
                messages.error(request, f'An error occurred while creating the client request: {str(e)}')
    else:
        form = ClientRequestForm(organization=organization)
    
    context = {
        'form': form,
        'title': 'New Client Request',
        'submit_text': 'Create Request'
    }
    
    return render(request, 'form_page.html', context)


def new_session(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                # Validate required fields
                required_fields = ['job_type', 'assigned_mechanic', 'scheduled_start', 'description']
                for field in required_fields:
                    if not request.POST.get(field):
                        messages.error(request, f'Please fill in the required field: {field.replace("_", " ").title()}')
                        return redirect('new_session')
                
                # Handle client creation/selection
                if request.POST.get('existing_client'):
                    client = Client.objects.get(id=request.POST['existing_client'], organization=organization)
                else:
                    # Validate new client fields
                    client_fields = ['client_first_name', 'client_last_name', 'client_phone']
                    for field in client_fields:
                        if not request.POST.get(field):
                            messages.error(request, f'Please select an existing client or fill in all new client fields')
                            return redirect('new_session')
                    
                    # Create new client
                    client = Client.objects.create(
                        organization=organization,
                        first_name=request.POST['client_first_name'],
                        last_name=request.POST['client_last_name'],
                        email=request.POST.get('client_email'),
                        phone=request.POST.get('client_phone')
                    )
                
                # Handle car creation/selection
                if request.POST.get('existing_car'):
                    car = Car.objects.get(id=request.POST['existing_car'], organization=organization)
                else:
                    # Validate new car fields
                    car_fields = ['car_make', 'car_model', 'car_year', 'car_license_plate']
                    for field in car_fields:
                        if not request.POST.get(field):
                            messages.error(request, f'Please select an existing vehicle or fill in all new vehicle fields')
                            return redirect('new_session')
                    
                    # Create new car
                    vin_value = request.POST.get('car_vin') or None  # Set to None if not provided
                    color_value = request.POST.get('car_color') or ''
                    
                    car = Car.objects.create(
                        organization=organization,
                        client=client,
                        make=request.POST['car_make'],
                        model=request.POST['car_model'],
                        year=request.POST['car_year'],
                        license_plate=request.POST['car_license_plate'],
                        vin=vin_value,
                        color=color_value
                    )
                
                # Create session
                session = Session.objects.create(
                    organization=organization,
                    car=car,
                    technician=UserProfile.objects.get(id=request.POST['assigned_mechanic'], organization=organization),
                    scheduled_date=request.POST['scheduled_start'],
                    expected_end_date=request.POST.get('expected_end'),
                    description=request.POST['description'],
                    status='scheduled',
                    job_type=request.POST['job_type']
                )
                
                # Create job cards dynamically
                job_card_count = 0
                while f'job_title_{job_card_count + 1}' in request.POST:
                    job_card_count += 1
                    JobCard.objects.create(
                        session=session,
                        title=request.POST[f'job_title_{job_card_count}'],
                        description=request.POST[f'job_description_{job_card_count}'],
                        priority=request.POST[f'job_priority_{job_card_count}'],
                        estimated_hours=request.POST.get(f'job_estimated_hours_{job_card_count}') or 0,
                        labor_cost=request.POST.get(f'job_labor_cost_{job_card_count}') or 0,
                        status=request.POST[f'job_status_{job_card_count}'],
                        assigned_technician=session.technician
                    )
                
                # Create parts usage dynamically
                part_count = 0
                while f'part_name_{part_count + 1}' in request.POST:
                    part_count += 1
                    part_name = request.POST[f'part_name_{part_count}']
                    if not part_name.strip():  # Skip empty part names
                        continue
                        
                    quantity = int(request.POST.get(f'part_quantity_{part_count}', 0))
                    unit_cost = float(request.POST.get(f'part_cost_{part_count}') or 0)
                    
                    if quantity <= 0:  # Skip invalid quantities
                        continue
                    
                    # Check if part exists in inventory
                    inventory_item = Inventory.objects.filter(
                        organization=organization,
                        name__iexact=part_name
                    ).first()
                    
                    if inventory_item:
                        # Update existing inventory (reduce quantity if sufficient stock)
                        if inventory_item.quantity >= quantity:
                            inventory_item.quantity -= quantity
                            inventory_item.save()
                        else:
                            # Log warning but continue - could be back-ordered parts
                            print(f"Warning: Insufficient stock for {part_name}. Available: {inventory_item.quantity}, Requested: {quantity}")
                    else:
                        # Create new inventory item with zero quantity (part not in stock but used)
                        # Generate unique part number
                        import uuid
                        part_number = f"PART-{uuid.uuid4().hex[:8].upper()}"
                        
                        # Ensure part number is unique
                        while Inventory.objects.filter(part_number=part_number, organization=organization).exists():
                            part_number = f"PART-{uuid.uuid4().hex[:8].upper()}"
                        
                        inventory_item = Inventory.objects.create(
                            organization=organization,
                            part_number=part_number,
                            name=part_name,
                            quantity=0,  # Start with 0 - part was used but not previously in inventory
                            unit_price=unit_cost,
                            category='other',  # Default category, can be updated later
                            min_quantity=1,  # Set low stock alert
                            description=f"Auto-created from session {session.session_number}"
                        )
                        print(f"Created new inventory item: {part_name} with 0 quantity (used but not in stock)")
                    
                    # Create inventory usage transaction record
                    InventoryTransaction.objects.create(
                        organization=organization,
                        inventory_item=inventory_item,
                        session=session,
                        transaction_type='usage',
                        quantity=quantity,
                        unit_price=unit_cost,
                        notes=f"Used in repair session {session.session_number}",
                        created_by=request.user
                    )
                
                # Handle file uploads
                if request.FILES.getlist('before_photos'):
                    # Handle before photos upload
                    pass  # Implement file handling as needed
                
                if request.FILES.getlist('before_videos'):
                    # Handle before videos upload
                    pass  # Implement file handling as needed
                
                messages.success(request, f'New repair session #{session.session_number} has been created successfully!')
                return redirect('repair_sessions')
                
        except Exception as e:
            messages.error(request, f'An error occurred while creating the session: {str(e)}')
            return redirect('new_session')
    
    else:
        # Get available clients, cars, and mechanics for the form
        clients = Client.objects.filter(organization=organization)
        cars = Car.objects.filter(organization=organization).select_related('client')
        mechanics = UserProfile.objects.filter(
            organization=organization,
            role__in=['technician', 'manager']
        ).select_related('user')
        
        context = {
            'clients': clients,
            'cars': cars,
            'mechanics': mechanics,
            'organization': organization,
            'title': 'Create New Repair Session'
        }
        
        return render(request, 'new_session.html', context)


def inspections(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    # Get all inspections for the organization
    inspections_list = Inspection.objects.filter(
        organization=organization
    ).select_related('car', 'car__client', 'inspector').order_by('-created_at')
    
    # Filter by status if provided
    status_filter = request.GET.get('status')
    if status_filter:
        inspections_list = inspections_list.filter(status=status_filter)
    
    # Search by inspection number or car
    search_query = request.GET.get('search')
    if search_query:
        inspections_list = inspections_list.filter(
            Q(inspection_number__icontains=search_query) |
            Q(car__make__icontains=search_query) |
            Q(car__model__icontains=search_query) |
            Q(car__license_plate__icontains=search_query) |
            Q(car__client__full_name__icontains=search_query)
        )
    
    context = {
        'inspections': inspections_list,
        'status_filter': status_filter,
        'search_query': search_query,
        'title': 'Inspections - Quality Control'
    }
    
    return render(request, 'inspections.html', context)


def new_inspection(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    if request.method == 'POST':
        form = InspectionForm(request.POST, organization=organization)
        if form.is_valid():
            inspection = form.save(commit=False)
            inspection.organization = organization
            inspection.inspector = user_profile
            inspection.save()
            
            messages.success(request, f'Inspection scheduled successfully! Inspection #{inspection.inspection_number} has been created.')
            return redirect('dashboard')
    else:
        form = InspectionForm(organization=organization)
    
    context = {
        'form': form,
        'title': 'Schedule New Inspection',
        'submit_text': 'Schedule Inspection'
    }
    
    return render(request, 'form_page.html', context)


def add_vehicle(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    if request.method == 'POST':
        form = VehicleForm(request.POST, organization=organization)
        if form.is_valid():
            vehicle = form.save(commit=False)
            vehicle.organization = organization
            vehicle.save()
            
            messages.success(request, f'Vehicle "{vehicle.make} {vehicle.model}" has been added successfully!')
            return redirect('dashboard')
    else:
        form = VehicleForm(organization=organization)
    
    context = {
        'form': form,
        'title': 'Add New Vehicle',
        'submit_text': 'Add Vehicle'
    }
    
    return render(request, 'form_page.html', context)


def add_inventory_item(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    if request.method == 'POST':
        form = InventoryItemForm(request.POST)
        if form.is_valid():
            inventory_item = form.save(commit=False)
            inventory_item.organization = organization
            inventory_item.save()
            
            messages.success(request, f'Inventory item "{inventory_item.name}" has been added successfully!')
            return redirect('inventory')
    else:
        form = InventoryItemForm()
    
    context = {
        'form': form,
        'title': 'Add Inventory Item',
        'submit_text': 'Add Item'
    }
    
    return render(request, 'form_page.html', context)


def view_inventory_item(request, item_id):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    try:
        item = Inventory.objects.get(id=item_id, organization=organization)
    except Inventory.DoesNotExist:
        messages.error(request, 'Inventory item not found.')
        return redirect('inventory')
    
    context = {
        'item': item,
        'title': f'View Inventory Item: {item.name}',
        'total_value': item.unit_price * item.quantity
    }
    
    return render(request, 'inventory_item_detail.html', context)


def edit_inventory_item(request, item_id):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    try:
        item = Inventory.objects.get(id=item_id, organization=organization)
    except Inventory.DoesNotExist:
        messages.error(request, 'Inventory item not found.')
        return redirect('inventory')
    
    if request.method == 'POST':
        form = InventoryItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, f'Inventory item "{item.name}" has been updated successfully!')
            return redirect('inventory')
    else:
        form = InventoryItemForm(instance=item)
    
    context = {
        'form': form,
        'item': item,
        'title': f'Edit Inventory Item: {item.name}',
        'submit_text': 'Update Item'
    }
    
    return render(request, 'form_page.html', context)


def view_session(request, session_id):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    try:
        session = Session.objects.get(id=session_id, organization=organization)
    except Session.DoesNotExist:
        messages.error(request, 'Session not found.')
        return redirect('appointments')
    
    # Get related job cards
    job_cards = session.job_cards.all().order_by('-created_at')
    
    context = {
        'session': session,
        'job_cards': job_cards,
        'title': f'View Session: {session.session_number}',
    }
    
    return render(request, 'session_detail.html', context)


def edit_session(request, session_id):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    try:
        session = Session.objects.get(id=session_id, organization=organization)
    except Session.DoesNotExist:
        messages.error(request, 'Session not found.')
        return redirect('appointments')
    
    if request.method == 'POST':
        form = SessionForm(request.POST, instance=session, organization=organization)
        if form.is_valid():
            form.save()
            messages.success(request, f'Session "{session.session_number}" has been updated successfully!')
            return redirect('appointments')
    else:
        form = SessionForm(instance=session, organization=organization)
    
    context = {
        'form': form,
        'session': session,
        'title': f'Edit Session: {session.session_number}',
        'submit_text': 'Update Session'
    }
    
    return render(request, 'form_page.html', context)


def start_session(request, session_id):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    try:
        session = Session.objects.get(id=session_id, organization=organization)
    except Session.DoesNotExist:
        messages.error(request, 'Session not found.')
        return redirect('repair_sessions')
    
    if request.method == 'POST':
        if session.status == 'scheduled':
            session.status = 'in_progress'
            session.actual_start_time = timezone.now()
            session.save()
            messages.success(request, f'Session "{session.session_number}" has been started successfully!')
        else:
            messages.warning(request, f'Session "{session.session_number}" cannot be started. Current status: {session.get_status_display()}')
        
        return redirect('view_session', session_id=session.id)
    
    context = {
        'session': session,
        'title': f'Start Session: {session.session_number}',
        'action': 'Start Session',
        'message': f'Are you sure you want to start session "{session.session_number}" for {session.car.make} {session.car.model}?',
        'confirm_text': 'Start Session'
    }
    
    return render(request, 'confirm_action.html', context)


def update_session_status(request, session_id):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    try:
        session = Session.objects.get(id=session_id, organization=organization)
    except Session.DoesNotExist:
        messages.error(request, 'Session not found.')
        return redirect('repair_sessions')
    
    if request.method == 'POST':
        new_status = request.POST.get('status')
        if new_status in dict(Session.STATUS_CHOICES):
            old_status = session.status
            session.status = new_status
            
            # Handle status-specific logic
            if new_status == 'in_progress' and not session.actual_start_time:
                session.actual_start_time = timezone.now()
            elif new_status == 'completed' and not session.actual_end_time:
                session.actual_end_time = timezone.now()
            
            session.save()
            messages.success(request, f'Session "{session.session_number}" status updated from {old_status} to {new_status}!')
        else:
            messages.error(request, 'Invalid status selected.')
        
        return redirect('view_session', session_id=session.id)
    
    context = {
        'session': session,
        'status_choices': Session.STATUS_CHOICES,
        'title': f'Update Status: {session.session_number}'
    }
    
    return render(request, 'update_session_status.html', context)


def complete_session(request, session_id):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    try:
        session = Session.objects.get(id=session_id, organization=organization)
    except Session.DoesNotExist:
        messages.error(request, 'Session not found.')
        return redirect('repair_sessions')
    
    if request.method == 'POST':
        # Calculate total cost from job cards
        total_cost = session.job_cards.aggregate(
            total=Sum(F('parts_cost') + F('labor_cost'))
        )['total'] or 0
        
        session.status = 'completed'
        session.actual_end_time = timezone.now()
        session.actual_cost = total_cost
        session.save()
        
        messages.success(request, f'Session "{session.session_number}" has been completed successfully! Total cost: ${total_cost:.2f}')
        return redirect('view_session', session_id=session.id)
    
    # Calculate estimated total cost
    estimated_cost = session.job_cards.aggregate(
        total=Sum(F('parts_cost') + F('labor_cost'))
    )['total'] or 0
    
    context = {
        'session': session,
        'estimated_cost': estimated_cost,
        'title': f'Complete Session: {session.session_number}',
        'action': 'Complete Session',
        'message': f'Are you sure you want to complete session "{session.session_number}"? This will calculate the final cost and mark the session as completed.',
        'confirm_text': 'Complete Session'
    }
    
    return render(request, 'confirm_action.html', context)


def generate_invoice(request, session_id):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        messages.error(request, 'User profile not found. Please contact administrator.')
        return redirect('home')
    
    try:
        session = Session.objects.get(id=session_id, organization=organization)
    except Session.DoesNotExist:
        messages.error(request, 'Session not found.')
        return redirect('repair_sessions')
    
    # Calculate totals
    job_cards = session.job_cards.all()
    total_parts_cost = job_cards.aggregate(total=Sum('parts_cost'))['total'] or 0
    total_labor_cost = job_cards.aggregate(total=Sum('labor_cost'))['total'] or 0
    total_cost = total_parts_cost + total_labor_cost
    
    # Create PDF invoice
    from reportlab.pdfgen import canvas
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from io import BytesIO
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    
    styles = getSampleStyleSheet()
    
    # Title
    title_style = styles['Heading1']
    title_style.alignment = 1  # Center
    elements.append(Paragraph(f"Invoice - Session {session.session_number}", title_style))
    elements.append(Spacer(1, 20))
    
    # Organization and Client info
    elements.append(Paragraph(f"<b>From:</b> {organization.name}", styles['Normal']))
    elements.append(Paragraph(f"<b>To:</b> {session.car.client.full_name}", styles['Normal']))
    elements.append(Paragraph(f"<b>Vehicle:</b> {session.car.make} {session.car.model} ({session.car.license_plate})", styles['Normal']))
    elements.append(Paragraph(f"<b>Date:</b> {timezone.now().strftime('%Y-%m-%d')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Job details
    elements.append(Paragraph("Work Performed:", styles['Heading2']))
    
    job_data = [['Job', 'Description', 'Parts Cost', 'Labor Cost', 'Total']]
    for job in job_cards:
        job_data.append([
            job.job_number,
            job.title,
            f'${job.parts_cost:.2f}',
            f'${job.labor_cost:.2f}',
            f'${(job.parts_cost + job.labor_cost):.2f}'
        ])
    
    job_table = Table(job_data)
    job_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(job_table)
    elements.append(Spacer(1, 20))
    
    # Totals
    elements.append(Paragraph(f"<b>Total Parts Cost:</b> ${total_parts_cost:.2f}", styles['Normal']))
    elements.append(Paragraph(f"<b>Total Labor Cost:</b> ${total_labor_cost:.2f}", styles['Normal']))
    elements.append(Paragraph(f"<b>Grand Total:</b> ${total_cost:.2f}", styles['Heading2']))
    
    doc.build(elements)
    
    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="invoice_{session.session_number}.pdf"'
    
    return response


def export_reports_csv(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        return redirect('home')
    
    # Get current date and time
    now = timezone.now()
    today = now.date()
    this_month = today.replace(day=1)
    this_year = today.replace(month=1, day=1)
    
    # Create the HttpResponse object with CSV header
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="gixat_reports_{today}.csv"'
    
    writer = csv.writer(response)
    
    # Write report header
    writer.writerow(['Gixat Auto Repair Reports'])
    writer.writerow(['Generated on:', now.strftime('%Y-%m-%d %H:%M:%S')])
    writer.writerow(['Organization:', organization.name])
    writer.writerow([])
    
    # Revenue Summary
    writer.writerow(['REVENUE SUMMARY'])
    writer.writerow(['Period', 'Amount'])
    
    revenue_today = Session.objects.filter(
        organization=organization,
        status='completed',
        actual_end_time__date=today
    ).aggregate(total=Sum('actual_cost'))['total'] or 0
    
    revenue_this_month = Session.objects.filter(
        organization=organization,
        status='completed',
        actual_end_time__date__gte=this_month
    ).aggregate(total=Sum('actual_cost'))['total'] or 0
    
    revenue_this_year = Session.objects.filter(
        organization=organization,
        status='completed',
        actual_end_time__date__gte=this_year
    ).aggregate(total=Sum('actual_cost'))['total'] or 0
    
    writer.writerow(['Today', f'${revenue_today:.2f}'])
    writer.writerow(['This Month', f'${revenue_this_month:.2f}'])
    writer.writerow(['This Year', f'${revenue_this_year:.2f}'])
    writer.writerow([])
    
    # Session Statistics
    writer.writerow(['SESSION STATISTICS'])
    total_sessions = Session.objects.filter(organization=organization).count()
    completed_sessions = Session.objects.filter(organization=organization, status='completed').count()
    completion_rate = (completed_sessions / total_sessions * 100) if total_sessions > 0 else 0
    
    writer.writerow(['Total Sessions', total_sessions])
    writer.writerow(['Completed Sessions', completed_sessions])
    writer.writerow(['Completion Rate', f'{completion_rate:.1f}%'])
    writer.writerow([])
    
    # Popular Services
    writer.writerow(['POPULAR SERVICES'])
    writer.writerow(['Service', 'Jobs Completed', 'Revenue'])
    
    popular_services = JobCard.objects.filter(
        session__organization=organization,
        status='completed'
    ).values('title').annotate(
        count=Count('id'),
        revenue=Sum(F('parts_cost') + F('labor_cost'))
    ).order_by('-count')[:10]
    
    for service in popular_services:
        writer.writerow([service['title'], service['count'], f'${service["revenue"]:.2f}'])
    
    writer.writerow([])
    
    # Technician Performance
    writer.writerow(['TECHNICIAN PERFORMANCE'])
    writer.writerow(['Technician', 'Sessions Completed', 'Total Revenue'])
    
    technician_performance = UserProfile.objects.filter(
        organization=organization,
        role__in=['technician', 'manager']
    ).annotate(
        sessions_completed=Count('session', filter=Q(session__status='completed')),
        total_revenue=Sum('session__actual_cost', filter=Q(session__status='completed'))
    ).order_by('-sessions_completed')
    
    for tech in technician_performance:
        writer.writerow([
            f'{tech.user.first_name} {tech.user.last_name}',
            tech.sessions_completed,
            f'${tech.total_revenue or 0:.2f}'
        ])
    
    writer.writerow([])
    
    # Monthly Revenue Trend
    writer.writerow(['MONTHLY REVENUE TREND (Last 12 Months)'])
    writer.writerow(['Month', 'Revenue'])
    
    for i in range(11, -1, -1):
        month_start = (now - timedelta(days=30*i)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        revenue = Session.objects.filter(
            organization=organization,
            status='completed',
            actual_end_time__date__gte=month_start,
            actual_end_time__date__lte=month_end
        ).aggregate(total=Sum('actual_cost'))['total'] or 0
        writer.writerow([month_start.strftime('%B %Y'), f'${revenue:.2f}'])
    
    return response


def export_reports_excel(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        return redirect('home')
    
    # Get current date and time
    now = timezone.now()
    today = now.date()
    this_month = today.replace(day=1)
    this_year = today.replace(month=1, day=1)
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # Revenue Summary Sheet
    ws_revenue = wb.active
    ws_revenue.title = "Revenue Summary"
    
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Title
    ws_revenue['A1'] = 'Gixat Auto Repair Reports'
    ws_revenue['A1'].font = Font(bold=True, size=16)
    ws_revenue['A2'] = f'Generated on: {now.strftime("%Y-%m-%d %H:%M:%S")}'
    ws_revenue['A3'] = f'Organization: {organization.name}'
    
    # Revenue data
    ws_revenue['A5'] = 'Period'
    ws_revenue['B5'] = 'Amount'
    ws_revenue['A5'].font = header_font
    ws_revenue['B5'].font = header_font
    ws_revenue['A5'].fill = header_fill
    ws_revenue['B5'].fill = header_fill
    
    revenue_today = Session.objects.filter(
        organization=organization,
        status='completed',
        actual_end_time__date=today
    ).aggregate(total=Sum('actual_cost'))['total'] or 0
    
    revenue_this_month = Session.objects.filter(
        organization=organization,
        status='completed',
        actual_end_time__date__gte=this_month
    ).aggregate(total=Sum('actual_cost'))['total'] or 0
    
    revenue_this_year = Session.objects.filter(
        organization=organization,
        status='completed',
        actual_end_time__date__gte=this_year
    ).aggregate(total=Sum('actual_cost'))['total'] or 0
    
    ws_revenue['A6'] = 'Today'
    ws_revenue['B6'] = revenue_today
    ws_revenue['B6'].number_format = '$#,##0.00'
    
    ws_revenue['A7'] = 'This Month'
    ws_revenue['B7'] = revenue_this_month
    ws_revenue['B7'].number_format = '$#,##0.00'
    
    ws_revenue['A8'] = 'This Year'
    ws_revenue['B8'] = revenue_this_year
    ws_revenue['B8'].number_format = '$#,##0.00'
    
    # Services Sheet
    ws_services = wb.create_sheet("Popular Services")
    ws_services['A1'] = 'Popular Services'
    ws_services['A1'].font = Font(bold=True, size=14)
    
    ws_services['A3'] = 'Service'
    ws_services['B3'] = 'Jobs Completed'
    ws_services['C3'] = 'Revenue'
    for cell in ['A3', 'B3', 'C3']:
        ws_services[cell].font = header_font
        ws_services[cell].fill = header_fill
    
    popular_services = JobCard.objects.filter(
        session__organization=organization,
        status='completed'
    ).values('title').annotate(
        count=Count('id'),
        revenue=Sum(F('parts_cost') + F('labor_cost'))
    ).order_by('-count')[:10]
    
    row = 4
    for service in popular_services:
        ws_services[f'A{row}'] = service['title']
        ws_services[f'B{row}'] = service['count']
        ws_services[f'C{row}'] = service['revenue']
        ws_services[f'C{row}'].number_format = '$#,##0.00'
        row += 1
    
    # Technicians Sheet
    ws_tech = wb.create_sheet("Technician Performance")
    ws_tech['A1'] = 'Technician Performance'
    ws_tech['A1'].font = Font(bold=True, size=14)
    
    ws_tech['A3'] = 'Technician'
    ws_tech['B3'] = 'Sessions Completed'
    ws_tech['C3'] = 'Total Revenue'
    for cell in ['A3', 'B3', 'C3']:
        ws_tech[cell].font = header_font
        ws_tech[cell].fill = header_fill
    
    technician_performance = UserProfile.objects.filter(
        organization=organization,
        role__in=['technician', 'manager']
    ).annotate(
        sessions_completed=Count('session', filter=Q(session__status='completed')),
        total_revenue=Sum('session__actual_cost', filter=Q(session__status='completed'))
    ).order_by('-sessions_completed')
    
    row = 4
    for tech in technician_performance:
        ws_tech[f'A{row}'] = f'{tech.user.first_name} {tech.user.last_name}'
        ws_tech[f'B{row}'] = tech.sessions_completed
        ws_tech[f'C{row}'] = tech.total_revenue or 0
        ws_tech[f'C{row}'].number_format = '$#,##0.00'
        row += 1
    
    # Monthly Trend Sheet
    ws_trend = wb.create_sheet("Monthly Revenue Trend")
    ws_trend['A1'] = 'Monthly Revenue Trend (Last 12 Months)'
    ws_trend['A1'].font = Font(bold=True, size=14)
    
    ws_trend['A3'] = 'Month'
    ws_trend['B3'] = 'Revenue'
    for cell in ['A3', 'B3']:
        ws_trend[cell].font = header_font
        ws_trend[cell].fill = header_fill
    
    row = 4
    for i in range(11, -1, -1):
        month_start = (now - timedelta(days=30*i)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        revenue = Session.objects.filter(
            organization=organization,
            status='completed',
            actual_end_time__date__gte=month_start,
            actual_end_time__date__lte=month_end
        ).aggregate(total=Sum('actual_cost'))['total'] or 0
        
        ws_trend[f'A{row}'] = month_start.strftime('%B %Y')
        ws_trend[f'B{row}'] = revenue
        ws_trend[f'B{row}'].number_format = '$#,##0.00'
        row += 1
    
    # Auto-adjust column widths
    for ws in [ws_revenue, ws_services, ws_tech, ws_trend]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="gixat_reports_{today}.xlsx"'
    
    wb.save(response)
    return response


def export_reports_pdf(request):
    if not request.user.is_authenticated:
        return redirect('login')
    
    try:
        user_profile = UserProfile.objects.get(user=request.user)
        organization = user_profile.organization
    except UserProfile.DoesNotExist:
        return redirect('home')
    
    # Get current date and time
    now = timezone.now()
    today = now.date()
    this_month = today.replace(day=1)
    this_year = today.replace(month=1, day=1)
    
    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="gixat_reports_{today}.pdf"'
    
    # Create PDF document
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=16,
        spaceAfter=20,
    )
    
    normal_style = styles['Normal']
    
    # Title
    elements.append(Paragraph("Gixat Auto Repair Reports", title_style))
    elements.append(Paragraph(f"Generated on: {now.strftime('%Y-%m-%d %H:%M:%S')}", normal_style))
    elements.append(Paragraph(f"Organization: {organization.name}", normal_style))
    elements.append(Spacer(1, 20))
    
    # Revenue Summary
    elements.append(Paragraph("Revenue Summary", subtitle_style))
    
    revenue_today = Session.objects.filter(
        organization=organization,
        status='completed',
        actual_end_time__date=today
    ).aggregate(total=Sum('actual_cost'))['total'] or 0
    
    revenue_this_month = Session.objects.filter(
        organization=organization,
        status='completed',
        actual_end_time__date__gte=this_month
    ).aggregate(total=Sum('actual_cost'))['total'] or 0
    
    revenue_this_year = Session.objects.filter(
        organization=organization,
        status='completed',
        actual_end_time__date__gte=this_year
    ).aggregate(total=Sum('actual_cost'))['total'] or 0
    
    revenue_data = [
        ['Period', 'Amount'],
        ['Today', f'${revenue_today:.2f}'],
        ['This Month', f'${revenue_this_month:.2f}'],
        ['This Year', f'${revenue_this_year:.2f}']
    ]
    
    revenue_table = Table(revenue_data)
    revenue_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(revenue_table)
    elements.append(Spacer(1, 20))
    
    # Session Statistics
    elements.append(Paragraph("Session Statistics", subtitle_style))
    
    total_sessions = Session.objects.filter(organization=organization).count()
    completed_sessions = Session.objects.filter(organization=organization, status='completed').count()
    completion_rate = (completed_sessions / total_sessions * 100) if total_sessions > 0 else 0
    
    stats_data = [
        ['Metric', 'Value'],
        ['Total Sessions', str(total_sessions)],
        ['Completed Sessions', str(completed_sessions)],
        ['Completion Rate', f'{completion_rate:.1f}%']
    ]
    
    stats_table = Table(stats_data)
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(stats_table)
    elements.append(Spacer(1, 20))
    
    # Popular Services
    elements.append(Paragraph("Popular Services", subtitle_style))
    
    popular_services = JobCard.objects.filter(
        session__organization=organization,
        status='completed'
    ).values('title').annotate(
        count=Count('id'),
        revenue=Sum(F('parts_cost') + F('labor_cost'))
    ).order_by('-count')[:10]
    
    services_data = [['Service', 'Jobs Completed', 'Revenue']]
    for service in popular_services:
        services_data.append([
            service['title'],
            str(service['count']),
            f'${service["revenue"]:.2f}'
        ])
    
    services_table = Table(services_data)
    services_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(services_table)
    elements.append(Spacer(1, 20))
    
    # Technician Performance
    elements.append(Paragraph("Technician Performance", subtitle_style))
    
    technician_performance = UserProfile.objects.filter(
        organization=organization,
        role__in=['technician', 'manager']
    ).annotate(
        sessions_completed=Count('session', filter=Q(session__status='completed')),
        total_revenue=Sum('session__actual_cost', filter=Q(session__status='completed'))
    ).order_by('-sessions_completed')
    
    tech_data = [['Technician', 'Sessions Completed', 'Total Revenue']]
    for tech in technician_performance:
        tech_data.append([
            f'{tech.user.first_name} {tech.user.last_name}',
            str(tech.sessions_completed),
            f'${tech.total_revenue or 0:.2f}'
        ])
    
    tech_table = Table(tech_data)
    tech_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(tech_table)
    
    # Monthly Trend Sheet
    ws_trend = wb.create_sheet("Monthly Revenue Trend")
    ws_trend['A1'] = 'Monthly Revenue Trend (Last 12 Months)'
    ws_trend['A1'].font = Font(bold=True, size=14)
    
    ws_trend['A3'] = 'Month'
    ws_trend['B3'] = 'Revenue'
    for cell in ['A3', 'B3']:
        ws_trend[cell].font = header_font
        ws_trend[cell].fill = header_fill
    
    row = 4
    for i in range(11, -1, -1):
        month_start = (now - timedelta(days=30*i)).replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        revenue = Session.objects.filter(
            organization=organization,
            status='completed',
            actual_end_time__date__gte=month_start,
            actual_end_time__date__lte=month_end
        ).aggregate(total=Sum('actual_cost'))['total'] or 0
        
        ws_trend[f'A{row}'] = month_start.strftime('%B %Y')
        ws_trend[f'B{row}'] = revenue
        ws_trend[f'B{row}'].number_format = '$#,##0.00'
        row += 1
    
    # Auto-adjust column widths
    for ws in [ws_revenue, ws_services, ws_tech, ws_trend]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="gixat_reports_{today}.pdf"'
    
    # Build PDF
    doc.build(elements)
    
    return response


class CustomLoginView(LoginView):
    template_name = 'login.html'
    form_class = CustomLoginForm
    
    def get_success_url(self):
        return reverse_lazy('dashboard')
    
    def form_valid(self, form):
        messages.success(self.request, f'Welcome back, {form.get_user().get_full_name()}!')
        return super().form_valid(form)

